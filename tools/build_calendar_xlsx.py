#!/usr/bin/env python3
"""
Собирает docs/calendar-data.json в читаемый Excel-файл для менеджмента.

Использование:
    python3 tools/build_calendar_xlsx.py
    python3 tools/build_calendar_xlsx.py --in docs/calendar-data.json --out docs/calendar-2026.xlsx
"""

import argparse
import json
import re
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

MONTHS_RU = ["", "январь", "февраль", "март", "апрель", "май", "июнь", "июль",
             "август", "сентябрь", "октябрь", "ноябрь", "декабрь"]

FONT_NAME = "Arial"
# Палитра приведена к дизайн-системе ГЭС-2 (design-system/design-system.md):
# сайт принципиально монохромный, фирменного цветного акцента нет — иерархия
# идёт через чёрный/белый и насыщенность, а не через отдельный брендовый цвет.
#
# Все цвета — 8-значный ARGB с явным альфа-каналом FF (непрозрачно). Это не
# опечатка: если передать openpyxl голый 6-значный RGB ("000000"), он молча
# дополняет строку нулями СЛЕВА, а не FF — получается 00000000 (альфа = 0,
# то есть прозрачно) вместо ожидаемого непрозрачного чёрного. Проверено
# эмпирически на этой версии openpyxl. Без явного FF шрифты и заливки рискуют
# оказаться невидимыми в Excel/LibreOffice.
INK = "FF000000"
ACCENT = "FF000000"       # раньше был зелёный "фирменный" акцент — заменён на чёрный
ACCENT_SOFT = "FFEEEEEE"  # подсветка вех — нейтральный hairline-серый, не цветной
MUTED = "FF6B6862"
LINE = "FFBFBFBF"
WHITE = "FFFFFFFF"
HEADER_FILL = PatternFill("solid", fgColor=ACCENT)
MILESTONE_FILL = PatternFill("solid", fgColor=ACCENT_SOFT)
STRIPE_FILL = PatternFill("solid", fgColor="FFF1EFE9")
THIN = Side(style="thin", color=LINE)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def clean(s):
    if s is None:
        return ""
    return re.sub(r"\s+", " ", s.replace("\xa0", " ")).strip()


def parse_dt(s):
    return datetime.fromisoformat(s) if s else None


def load_events(path):
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    for e in data:
        e["starts_dt"] = parse_dt(e["starts"])
        e["ends_dt"] = parse_dt(e["ends"])
        e["subtype"] = clean(e["subtype"])
        e["title"] = clean(e["title"])
    data.sort(key=lambda e: e["starts_dt"])
    return data


def build(events, out_path):
    wb = Workbook()

    # ---------- Сводка ----------
    ws = wb.active
    ws.title = "Сводка"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 46

    ws["A1"] = "Календарь мероприятий ГЭС-2 — до конца 2026"
    ws["A1"].font = Font(name=FONT_NAME, size=16, bold=True, color=INK)
    ws.merge_cells("A1:B1")

    ws["A2"] = "Источник: ges-2.org/graphql (operation newMaterialsInterval) · собрано 09.08.2026"
    ws["A2"].font = Font(name=FONT_NAME, size=10, italic=True, color=MUTED)
    ws.merge_cells("A2:B2")

    row = 4
    ws.cell(row=row, column=1, value="Показатель").font = Font(name=FONT_NAME, bold=True, color=WHITE)
    ws.cell(row=row, column=2, value="Значение").font = Font(name=FONT_NAME, bold=True, color=WHITE)
    ws.cell(row=row, column=1).fill = HEADER_FILL
    ws.cell(row=row, column=2).fill = HEADER_FILL

    # Значения посчитаны в Python на момент сборки файла, а не формулами Excel:
    # в этом окружении не установлен LibreOffice, а без него нельзя проверить,
    # что формулы openpyxl действительно вычислятся без ошибок при открытии.
    # Пересборка файла (см. tools/build_calendar_xlsx.py) обновит и эти цифры.
    subtypes = [e["subtype"] for e in events]
    stats = [
        ("Всего событий в диапазоне", len(events)),
        ("Из них платные", sum(1 for e in events if e["paid"])),
        ("Выставок (записей типа «Выставка»)", subtypes.count("Выставка")),
        ("Кинопоказов", subtypes.count("Кинопоказ")),
        ("Медиаторских туров", subtypes.count("Медиаторский тур")),
        ("Мастер-классов", subtypes.count("Мастер-класс")),
    ]
    for i, (label, value) in enumerate(stats, start=row + 1):
        ws.cell(row=i, column=1, value=label).font = Font(name=FONT_NAME, size=11, color=INK)
        c = ws.cell(row=i, column=2, value=value)
        c.font = Font(name=FONT_NAME, size=11, color=INK, bold=True)
        c.alignment = Alignment(horizontal="center")
        fill = STRIPE_FILL if i % 2 == 0 else PatternFill(fill_type=None)
        ws.cell(row=i, column=1).fill = fill
        ws.cell(row=i, column=2).fill = fill

    note_row = row + len(stats) + 2
    ws.cell(row=note_row, column=1, value="Ключевые операционные точки").font = Font(name=FONT_NAME, size=12, bold=True, color=ACCENT)
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=2)

    notes = [
        "Три текущие выставки закрываются почти одновременно: Тагор — 23.08, Хадзар мечты — 13.09, Москва — Ханой — 20.09.",
        "Следующая выставка «Натюрморт. Портрет. Пейзаж. После жанра» открывается только 15.10 — окно ~месяц на монтаж/демонтаж.",
        "«ГЭС-2 Саундсистем 4» (24–30.08) — крупный музыкальный уикенд, пиковая нагрузка на посетительский сервис и парковку.",
        "Горизонт публикации сайта — ~3–4 месяца: событий после 10.12.2026 сайт пока не публикует.",
        "Обновить данные: python3 tools/fetch_calendar.py --from 2026-08-09 --to 2026-12-31 --out docs/calendar-data.json",
        "Цифры на этом листе — статические значения, посчитанные при сборке файла (tools/build_calendar_xlsx.py), а не live-формулы Excel: в окружении, где собирался файл, недоступен LibreOffice для проверки пересчёта.",
    ]
    for i, n in enumerate(notes, start=note_row + 1):
        c = ws.cell(row=i, column=1, value="•  " + n)
        c.font = Font(name=FONT_NAME, size=10.5, color=INK)
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=2)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[i].height = 28

    # ---------- Календарь ----------
    ws2 = wb.create_sheet("Календарь")
    headers = ["Дата начала", "Дата окончания", "Месяц", "Подтип", "Тип материала", "Название", "Платно", "Ссылка"]
    widths = [14, 16, 14, 22, 16, 55, 9, 44]
    for col, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(name=FONT_NAME, bold=True, color=WHITE)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.freeze_panes = "A2"
    ws2.row_dimensions[1].height = 20

    exhibition_project_ids = {e["id"] for e in events if e["subtype"] == "Выставка" and e["type"] == "project"}

    for r, e in enumerate(events, start=2):
        s = e["starts_dt"]
        en = e["ends_dt"]
        month_label = f"{MONTHS_RU[s.month]} {s.year}" if s else ""
        material_type = "Выставка/проект" if e["type"] == "project" else "Событие"
        values = [
            s.replace(tzinfo=None) if s else None,
            en.replace(tzinfo=None) if en else None,
            month_label,
            e["subtype"],
            material_type,
            e["title"],
            "Да" if e["paid"] else "Нет",
            f"https://ges-2.org/{e['path']}",
        ]
        is_milestone = e["id"] in exhibition_project_ids
        for col, v in enumerate(values, start=1):
            cell = ws2.cell(row=r, column=col, value=v)
            cell.font = Font(name=FONT_NAME, size=10.5, color=ACCENT if is_milestone else INK,
                              bold=is_milestone)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=(col == 6))
            if col in (1, 2):
                cell.number_format = "dd.mm.yyyy"
            if is_milestone:
                cell.fill = MILESTONE_FILL
            elif r % 2 == 0:
                cell.fill = STRIPE_FILL

    last_row = len(events) + 1
    table = Table(displayName="Calendar2026", ref=f"A1:H{last_row}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showRowStripes=False,
                                           showFirstColumn=False, showLastColumn=False)
    ws2.add_table(table)

    wb.save(out_path)
    print(f"Записано {len(events)} строк в {out_path}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--in", dest="in_path", default="docs/calendar-data.json")
    ap.add_argument("--out", dest="out_path", default="docs/calendar-2026.xlsx")
    args = ap.parse_args()
    events = load_events(args.in_path)
    build(events, args.out_path)


if __name__ == "__main__":
    main()
