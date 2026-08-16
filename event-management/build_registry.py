#!/usr/bin/env python3
"""
Собирает автономный index.html системы управления событиями ГЭС-2
из docs/ges2_event_pm_template.xlsx.

Извлекает три листа:
  - "Реестр событий"    -> список событий (72 записи на 09.08.2026)
  - "Этапы по трекам"   -> справочник этапов по трекам сложности (T-минус окна)
  - "RACI"              -> матрица ответственности по трекам/этапам/отделам

"Текущий этап" в исходном файле — статичное значение (не формула), явно
описанное в листе "Легенда" как "стартовая подсказка, не подтверждённый
факт". Чтобы инструмент оставался полезным по мере того, как проходит
время (а не застыл на дате сборки), мы НЕ копируем это статичное значение,
а пересчитываем этап на лету в браузере по текущей дате посетителя —
см. computeStage() в template.html. Формула восстановлена и проверена
эмпирически по исходным данным: для всех 23 событий, которые идут "сейчас"
(today между датой начала и датой окончания), исходное значение было
"Проведение" независимо от трека и T-минус; для будущих событий исходное
значение точно соответствует окну "Старт..Дедлайн" из листа "Этапы по
трекам" (по модулю), с откатом на самый ранний этап, если событие дальше
горизонта планирования трека.

Использование:
    python3 build_registry.py
    python3 build_registry.py --xlsx ../docs/ges2_event_pm_template.xlsx --out index.html
"""

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

HERE = Path(__file__).parent

CATEGORIES = ["Концерт", "Танцевальный показ", "Медиаторский тур", "Выставка", "Кинопоказ", "Лекция", "Прочее"]
TRACKS = ["Лёгкий", "Средний", "Тяжёлый"]
STAGES_ORDER = [
    "Заявка/утверждение", "Бронирование площадки и ресурсов", "Продакшн",
    "Маркетинг/PR-план", "Финальная неделя", "Проведение", "Пост-ивент",
]
DEPARTMENTS = ["Продакшн/техника", "Маркетинг", "PR", "Площадка", "Питание", "Охрана"]
STATUS_VALUES = ["Не начат", "Зелёный", "Жёлтый", "Красный", "Н/Д"]

REGISTRY_DEPT_COLS = {
    "Продакшн/техника": "Продакшн/техника",
    "Маркетинг": "Маркетинг",
    "PR": "PR",
    "Площадка": "Площадка",
    "Питание": "Питание",
    "Охрана": "Охрана",
}


def iso(dt):
    return dt.strftime("%Y-%m-%d") if dt else None


def extract_registry(ws):
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}
    events = []
    errors = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[idx["Название события"]] is None:
            continue
        eid = row[idx["ID"]]
        category = row[idx["Категория"]]
        track = row[idx["Трек"]]
        date_start = row[idx["Дата начала"]]
        date_end = row[idx["Дата окончания"]]

        if category not in CATEGORIES:
            errors.append(f"ID {eid}: неизвестная категория '{category}'")
        if track not in TRACKS:
            errors.append(f"ID {eid}: неизвестный трек '{track}'")
        if not isinstance(date_start, datetime) or not isinstance(date_end, datetime):
            errors.append(f"ID {eid}: не заданы даты начала/окончания")
            continue

        departments = {}
        for dept, col in REGISTRY_DEPT_COLS.items():
            val = row[idx[col]]
            if val is not None and val not in STATUS_VALUES:
                errors.append(f"ID {eid}: недопустимый статус '{val}' в столбце '{col}'")
            departments[dept] = val

        events.append({
            "id": eid,
            "title": row[idx["Название события"]],
            "category": category,
            "track": track,
            "dateStart": iso(date_start),
            "dateEnd": iso(date_end),
            "producer": row[idx["Продюсер"]],
            "sourceStageHint": row[idx["Текущий этап"]],
            "departments": departments,
            "paid": row[idx["Платно"]] == "Да",
            "riskNote": row[idx["Риски / комментарий"]],
            "link": row[idx["Ссылка"]],
        })

    return events, errors


def extract_stages(ws):
    by_track = {t: [] for t in TRACKS}
    for row in ws.iter_rows(min_row=2, values_only=True):
        track = row[0]
        if track not in TRACKS:
            continue
        by_track[track].append({
            "stage": row[1],
            "startTMinus": row[2],
            "deadlineTMinus": row[3],
            "responsible": row[4],
            "comment": row[5],
        })
    return by_track


def extract_raci(ws):
    """RACI-лист организован тремя блоками ('Трек: <название>'), в каждом —
    строка-заголовок с отделами, затем строки этапов с R/A/C/I."""
    raci = {t: {} for t in TRACKS}
    current_track = None
    dept_cols = None
    for row in ws.iter_rows(min_row=1, values_only=True):
        first = row[0]
        if isinstance(first, str) and first.startswith("Трек: "):
            current_track = first.replace("Трек: ", "").strip()
            dept_cols = None
            continue
        if first == "Этап":
            dept_cols = row[1:]
            continue
        if current_track and dept_cols and first in STAGES_ORDER:
            raci[current_track][first] = {
                dept: role for dept, role in zip(dept_cols, row[1:]) if dept
            }
    return raci


def validate(events, errors):
    ids = [e["id"] for e in events]
    if len(ids) != len(set(ids)):
        errors.append("обнаружены дублирующиеся ID в реестре")
    if errors:
        print("Сборка остановлена — данные не прошли проверку:", file=sys.stderr)
        for err in errors:
            print(f"  - {err}", file=sys.stderr)
        sys.exit(1)


def build(xlsx_path: Path, template_path: Path, out_path: Path, data_out_path: Path):
    wb = load_workbook(xlsx_path, data_only=True)

    events, errors = extract_registry(wb["Реестр событий"])
    stages_by_track = extract_stages(wb["Этапы по трекам"])
    raci = extract_raci(wb["RACI"])
    validate(events, errors)

    data = {
        "meta": {
            "generated": "2026-08-09",
            "source": "docs/ges2_event_pm_template.xlsx",
            "totalEvents": len(events),
        },
        "categories": CATEGORIES,
        "tracks": TRACKS,
        "stagesOrder": STAGES_ORDER,
        "departments": DEPARTMENTS,
        "statusValues": STATUS_VALUES,
        "stagesByTrack": stages_by_track,
        "raci": raci,
        "events": events,
    }

    data_out_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

    data_json = json.dumps(data, ensure_ascii=False).replace("</", "<\\/")
    template = template_path.read_text(encoding="utf-8")
    html = (
        template
        .replace("__REGISTRY_DATA_JSON__", data_json)
        .replace("__GENERATED_DATE__", data["meta"]["generated"])
        .replace("__TOTAL__", str(len(events)))
    )
    out_path.write_text(html, encoding="utf-8")

    print(f"OK: {data_out_path} ({len(events)} событий, справочник этапов и RACI)")
    print(f"OK: {out_path} собран")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=str(HERE / ".." / "docs" / "ges2_event_pm_template.xlsx"))
    ap.add_argument("--template", default=str(HERE / "template.html"))
    ap.add_argument("--out", default=str(HERE / "index.html"))
    ap.add_argument("--data-out", default=str(HERE / "registry-data.json"))
    args = ap.parse_args()
    build(Path(args.xlsx), Path(args.template), Path(args.out), Path(args.data_out))


if __name__ == "__main__":
    main()
